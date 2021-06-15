import warnings
from pandas.core.common import SettingWithCopyWarning

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from datetime import date
import pandas as pd
from tabulate import tabulate

warnings.simplefilter(action='ignore', category=SettingWithCopyWarning)

AemFile = 'C:/Users/Georges/Downloads/AEM-Emails.txt'
file = open(AemFile, 'r')

today = date.today()
myToday = str(today).replace('-', '')

year = date.today().year

listSource = []
listTitle = []
listFirstname = []
listLastname = []
listEmail = []
listSpeciality = []
listCountry = []
listPhone = []
listCompany = []

listTheme = []
listType = []

lines = file.readlines()

for num, x in enumerate(lines):
    if x == 'Subject:\tAMS Virtual - Contact Us\n':
        listSource.append('AMS FACE '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tFACE 2021 Program Download - Delegate Prospect Lead\n':
        listSource.append('AEM FACE '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tContact us - FACE\n':
        listSource.append('AEM FACE '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tFACE exhibitor pack Download - Lead\n':
        listSource.append('AEM FACE '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+24].replace('Landline (business) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append(lines[num+18].replace('Organization Alias \t', '').replace('\n', '').replace(' \t', '').title())
        listTheme.append('AA')
        listType.append('EXHIBITOR')

    if x == 'Subject:\tVISAGE 2021 Exhibitor pack Download - Lead\n':
        listSource.append('AEM VISAGE '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+24].replace('Landline (business) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append(lines[num+18].replace('Organization Alias \t', '').replace('\n', '').replace(' \t', '').title())
        listTheme.append('AA')
        listType.append('EXHIBITOR')

    if x == 'Subject:\tVISAGE 2021 - I would like to be contacted for Sponsorship - Leads\n':
        listSource.append('AEM VISAGE '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+24].replace('Landline (business) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append(lines[num+18].replace('Organization Alias \t', '').replace('\n', '').replace(' \t', '').title())
        listTheme.append('AA')
        listType.append('EXHIBITOR')

    if x == 'Subject:\tEUROGIN - Webinar Re-watch Contacts\n':
        listSource.append('AEM EUROGIN '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('GYN')
        listType.append('DELEGATE')

    if x == 'Subject:\tEUROGIN Contact Us 2021\n':
        listSource.append('AEM EUROGIN '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('GYN')
        listType.append('DELEGATE')

    if x == 'Subject:\tInscription au Workshop Francophone\n':
        listSource.append('AEM EUROGIN FRENCH-WS '+str(year))
        listTitle.append('')
        listFirstname.append(lines[num+6].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+7].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+8].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append('Other')
        listCountry.append(lines[num+9].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('GYN')
        listType.append('DELEGATE')

    if x == 'Subject:\tFrancophone Workshop - Inscription\n':
        listSource.append('AEM EUROGIN FRENCH-WS '+str(year))
        listTitle.append('')
        listFirstname.append(lines[num+6].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+7].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+8].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append('Other')
        listCountry.append(lines[num+9].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('GYN')
        listType.append('DELEGATE')

    if x == 'Subject:\tEUROGIN 2021 Sponsorship Brochure Download - Exhibitor\n':
        listSource.append('AEM EUROGIN '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+24].replace('Landline (business) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append(lines[num+18].replace('Organization Alias \t', '').replace('\n', '').replace(' \t', '').title())
        listTheme.append('GYN')
        listType.append('EXHIBITOR')

    if x == 'Subject:\tAMWC Asia 2021 Program Download - Delegate Prospect Lead\n':
        listSource.append('AEM AMWC-ASIA '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tAMWC 2021 - I would like to be contacted - Exhibitor\n':
        listSource.append('AEM AMWC '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+24].replace('Mobile (personal) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append(lines[num+18].replace('Organization Alias \t', '').replace('\n', '').replace(' \t', '').title())
        listTheme.append('AA')
        listType.append('EXHIBITOR')

    if x == 'Subject:\tAMWC 2021 Newsletter Subscription\n':
        listSource.append('AEM AMWC '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+24].replace('Mobile (personal) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append(lines[num+18].replace('Organization Alias \t', '').replace('\n', '').replace(' \t', '').title())
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tAMWC Monaco 2021 - Contact Us\n':
        listSource.append('AEM AMWC '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tAMWC Monaco Program Download - Delegate Prospect Lead\n':
        listSource.append('AEM AMWC '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tAMWC Global Contest - Fotona\n':
        listSource.append('AEM AMWC '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(business) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append('Other')
        listCountry.append(lines[num+10].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append('')
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')

    if x == 'Subject:\tICAD Bangkok 2021 - I would like to be contacted - Exhibitor\n':
        listSource.append('AEM ICAD '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+24].replace('Mobile (personal) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append(lines[num+18].replace('Organization Alias \t', '').replace('\n', '').replace(' \t', '').title())
        listTheme.append('AA')
        listType.append('EXHIBITOR')

    if x == 'Subject:\tICAD Bangkok 2021 - I would like to attend - Delegate Prospect\n':
        listSource.append('AEM ICAD '+str(year))
        listTitle.append(lines[num+6].replace('Title \t', '').replace('\n', '').replace(' \t', ''))
        listFirstname.append(lines[num+7].replace('First Name \t', '').replace('\n', '').replace(' \t', '').title())
        listLastname.append(lines[num+8].replace('Surname \t', '').replace('\n', '').replace(' \t', '').title())
        listEmail.append(lines[num+9].replace('Email Address(personal) \t', '').replace('\n', '').replace(' \t', '').lower())
        listSpeciality.append(lines[num+10].replace('Job Title \t', '').replace('\n', '').replace(' \t', '').title())
        listCountry.append(lines[num+11].replace('Country \t', '').replace('\n', '').replace(' \t', '')[:-3])
        listPhone.append(lines[num+20].replace('Mobile (personal) \t', '').replace('\n', '').replace(' \t', ''))
        listCompany.append('')
        listTheme.append('AA')
        listType.append('DELEGATE')


MergeLists = list(zip(listSource, listTheme, listType, listTitle, listFirstname, listLastname, listEmail, listSpeciality, listCountry, listPhone, listCompany))

df = pd.DataFrame(MergeLists, columns=['source', 'theme', 'type', 'title', 'firstname', 'lastname', 'email', 'speciality', 'country', 'phone', 'company'])

# FIX TITLE
df['title'] = df['title'].replace(['Dr. (DR.)'], 'Dr')
df['title'] = df['title'].replace(['Ms. (MS.)'], 'Ms')
df['title'] = df['title'].replace(['Mrs. (MRS.)'], 'Mrs')
df['title'] = df['title'].replace(['Mr. (MR.)'], 'Mr')
df['title'] = df['title'].replace(['Prof. (PROF)'], 'Prof.')
df['title'] = df['title'].replace(['Other (OTHER)'], '')


# FIX SPECIALITY
df['speciality'] = df['speciality'].replace({' And ': ' & '}, regex=True)

df['speciality'] = df['speciality'].replace(['Dentistry'], 'Dentist')

df.loc[df['speciality'].str.contains('Cardiolo'), 'speciality'] = 'Cardiology'

df.loc[df['speciality'].str.contains('Dermatolog'), 'speciality'] = 'Dermatologist'
df['speciality'] = df['speciality'].replace(['Skin Therapist'], 'Dermatologist')
df['speciality'] = df['speciality'].replace(['Dernatology'], 'Dermatologist')
df['speciality'] = df['speciality'].replace(['Dermatolgy'], 'Dermatologist')

df['speciality'] = df['speciality'].replace(['Cosmetologist'], 'Cosmetologist MD')
df['speciality'] = df['speciality'].replace(['Cosmétologue'], 'Cosmetologist MD')

df.loc[df['speciality'].str.contains('Research'), 'speciality'] = 'Research Scientist'
df['speciality'] = df['speciality'].replace(['Scienctific Field'], 'Research Scientist')
df['speciality'] = df['speciality'].replace(['Scientific Field'], 'Research Scientist')

df.loc[df['speciality'].str.contains('Education'), 'speciality'] = 'Professor/Teacher'
df['speciality'] = df['speciality'].replace(['Pr'], 'Professor/Teacher')
df['speciality'] = df['speciality'].replace(['Professor'], 'Professor/Teacher')
df['speciality'] = df['speciality'].replace(['Teacher'], 'Professor/Teacher')

df['speciality'] = df['speciality'].replace(['Phd- Biomedicine'], 'Biomedical Physician')
df['speciality'] = df['speciality'].replace(['Phd - Biomedicine'], 'Biomedical Physician')
df['speciality'] = df['speciality'].replace(['Phd Biomedicine'], 'Biomedical Physician')
df['speciality'] = df['speciality'].replace(['Molecular Biology'], 'Biomedical Physician')

df.loc[df['speciality'].str.contains('Microbio'), 'speciality'] = 'Biologist/Biochemist'
df['speciality'] = df['speciality'].replace(['Biology'], 'Biologist/Biochemist')
df['speciality'] = df['speciality'].replace(['Biologist'], 'Biologist/Biochemist')
df['speciality'] = df['speciality'].replace(['Biochemist'], 'Biologist/Biochemist')

df.loc[df['speciality'].str.contains('Ophthalmo'), 'speciality'] = 'Ophthalmologist'

df.loc[df['speciality'].str.contains('Medical Device'), 'speciality'] = 'Surgical Equipment'

df['speciality'] = df['speciality'].replace(['Craniofacial Surgery Cosmetology'], 'Maxillofacial Surgeon')

df['speciality'] = df['speciality'].replace(['Plastic'], 'Aesthetic & Plastic Surgeon')
df['speciality'] = df['speciality'].replace(['Plastic Surgery'], 'Aesthetic & Plastic Surgeon')
df['speciality'] = df['speciality'].replace(['Aesthetics Surgery'], 'Aesthetic & Plastic Surgeon')
df['speciality'] = df['speciality'].replace(['Plastic Surgen'], 'Aesthetic & Plastic Surgeon')

df['speciality'] = df['speciality'].replace(['Surgery'], 'General Surgeon')
df['speciality'] = df['speciality'].replace(['Surgeon'], 'General Surgeon')

df['speciality'] = df['speciality'].replace(['Md, Mph'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Md'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Doctor'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['General'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Gp'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Family Medicine'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Médico'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Medicine'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Medical Education'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['Physician'], 'General Practitioner')
df['speciality'] = df['speciality'].replace(['General Physician'], 'General Practitioner')

df['speciality'] = df['speciality'].replace(['Urology'], 'Urologist')

df['speciality'] = df['speciality'].replace(['Epidemiology'], 'Epidemiologist')

df.loc[df['speciality'].str.contains('Anesthesist'), 'speciality'] = 'Anesthesist'
df['speciality'] = df['speciality'].replace(['Anaesthetist'], 'Anesthesist')
df['speciality'] = df['speciality'].replace(['Anesthesiology'], 'Anesthesist')

df['speciality'] = df['speciality'].replace(['Podiatry'], 'Pediatrist')

df['speciality'] = df['speciality'].replace(['Oculoplastics'], 'Oculoplastic Surgeon')

df['speciality'] = df['speciality'].replace(['Cosmetic Surgery'], 'Cosmetic Surgeon')

df['speciality'] = df['speciality'].replace(['Facial Plastics'], 'Facial Plastic Surgeon')

df['speciality'] = df['speciality'].replace(['Rehabilitation Medicine'], 'Plastic & Reconstructive Surgeon')

df.loc[df['speciality'].str.contains('Oncolog'), 'speciality'] = 'Oncologist'
df['speciality'] = df['speciality'].replace(['Infection & Cancer'], 'Oncologist')

df.loc[df['speciality'].str.contains('Cytolo'), 'speciality'] = 'Cytologist'
df['speciality'] = df['speciality'].replace(['Cytotechnologist'], 'Cytologist')

df['speciality'] = df['speciality'].replace(['Hpv'], 'Virologist')
df['speciality'] = df['speciality'].replace(['Virology'], 'Virologist')

df['speciality'] = df['speciality'].replace(['Neurology & Aesthetics'], 'Neurologist')

df.loc[df['speciality'].str.contains('Hair Transplant'), 'speciality'] = 'Hair Transplant'
df['speciality'] = df['speciality'].replace(['Hair'], 'Hair Transplant')

df.loc[df['speciality'].str.contains('Ob/Gy'), 'speciality'] = 'Gynecologist/Obstetrician'
df.loc[df['speciality'].str.contains('Gynaecolo'), 'speciality'] = 'Gynecologist/Obstetrician'
df.loc[df['speciality'].str.contains('Gynecolo'), 'speciality'] = 'Gynecologist/Obstetrician'
df.loc[df['speciality'].str.contains('Ginecolo'), 'speciality'] = 'Gynecologist/Obstetrician'
df.loc[df['speciality'].str.contains('Gynekolo'), 'speciality'] = 'Gynecologist/Obstetrician'
df.loc[df['speciality'].str.contains('Obstetrics'), 'speciality'] = 'Gynecologist/Obstetrician'
df['speciality'] = df['speciality'].replace(['Gynäkologie'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Gynae'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Gyn'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Ob Gyn'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Ob-Gyn'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Gyn/Ob'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Obgyn'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Ob &Gyn'], 'Gynecologist/Obstetrician')
df['speciality'] = df['speciality'].replace(['Ob&Gyn'], 'Gynecologist/Obstetrician')

df.loc[df['speciality'].str.contains('Nurse Aesthetic'), 'speciality'] = 'Aesthetic Practitioner'
df.loc[df['speciality'].str.contains('Medical Aesthetic'), 'speciality'] = 'Aesthetic Practitioner'
df['speciality'] = df['speciality'].replace(['Médecine Esthétique'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetic'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aethetic'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetics'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aeathetic'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetic Physiscian'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetic Complications'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aestheticx'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetic Medicine'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Ästhetische Medizin'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetic Nurse Prescriber'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetic Doctor'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Aesthetic Nurse Practitioner'], 'Aesthetic Practitioner')
df['speciality'] = df['speciality'].replace(['Cosmetic Docter'], 'Aesthetic Practitioner')

df.loc[df['speciality'].str.contains('Head & Neck'), 'speciality'] = 'ENT/Head & Neck Specialist'
df.loc[df['speciality'].str.contains('Otorhinolaryngo'), 'speciality'] = 'ENT/Head & Neck Specialist'
df.loc[df['speciality'].str.contains('Otolaryngo'), 'speciality'] = 'ENT/Head & Neck Specialist'
df['speciality'] = df['speciality'].replace(['Ent'], 'ENT/Head & Neck Specialist')
df['speciality'] = df['speciality'].replace(['Ent/ Head And Neck Surgery'], 'ENT/Head & Neck Specialist')

df.loc[df['speciality'].str.contains('Immunolo'), 'speciality'] = 'Immunologist'
df.loc[df['speciality'].str.contains('Vaccines'), 'speciality'] = 'Immunologist'

df.loc[df['speciality'].str.contains('Androlo'), 'speciality'] = 'Andrologist'
df['speciality'] = df['speciality'].replace(['Dsds'], 'Andrologist')

df.loc[df['speciality'].str.contains('Public Health'), 'speciality'] = 'Public Health Specialist'

df['speciality'] = df['speciality'].replace(['Advance Practice Nurse Practitioner, Owner Of Happy Body Atx'], 'Nurse (Registered - ARNP)')

df['speciality'] = df['speciality'].replace(['Nursing'], 'Nurse')
df['speciality'] = df['speciality'].replace(['Nurse Prescriber'], 'Nurse')
df['speciality'] = df['speciality'].replace(['Midwife'], 'Nurse')

df.loc[df['speciality'].str.contains('Pharmacist'), 'speciality'] = 'Pharmacist'
df['speciality'] = df['speciality'].replace(['Pharma'], 'Pharmacist')

df.loc[df['speciality'].str.contains('Pathology'), 'speciality'] = 'Pathologist'

df['speciality'] = df['speciality'].replace(['Industry'], 'Medical Aesthetics Industry')

df['speciality'] = df['speciality'].replace(['Sponsor'], 'Event Manager')
df['speciality'] = df['speciality'].replace(['Exhibitor'], 'Event Manager')
df['speciality'] = df['speciality'].replace(['Congress Manager'], 'Event Manager')
df['speciality'] = df['speciality'].replace(['Event & Marketing Coordinator'], 'Event Manager')

df['speciality'] = df['speciality'].replace(['Journalism'], 'Press')
df['speciality'] = df['speciality'].replace(['Publisher'], 'Press')

df['speciality'] = df['speciality'].replace(['Clinic Owner'], 'Clinic Manager')

df.loc[df['speciality'].str.contains('Marketing'), 'speciality'] = 'Marketing Manager'

df.loc[df['speciality'].str.contains('Medical Affair'), 'speciality'] = 'Business Development'
df.loc[df['speciality'].str.contains('Development'), 'speciality'] = 'Business Development'
df.loc[df['speciality'].str.contains('Business'), 'speciality'] = 'Business Development'

df.loc[df['speciality'].str.contains('Consulting'), 'speciality'] = 'Consultant'

df.loc[df['speciality'].str.contains('Area Manager'), 'speciality'] = 'General Manager'
df['speciality'] = df['speciality'].replace(['Présidente'], 'General Manager')
df['speciality'] = df['speciality'].replace(['President'], 'General Manager')

df.loc[df['speciality'].str.contains('Founder'), 'speciality'] = 'Company Manager'
df.loc[df['speciality'].str.contains('Director'), 'speciality'] = 'Company Manager'
df.loc[df['speciality'].str.contains('Directeur'), 'speciality'] = 'Company Manager'
df['speciality'] = df['speciality'].replace(['Country Manager'], 'Company Manager')
df['speciality'] = df['speciality'].replace(['Regional Manager'], 'Company Manager')
df['speciality'] = df['speciality'].replace(['Managing Director'], 'Company Manager')
df['speciality'] = df['speciality'].replace(['Head Of Strategy'], 'Company Manager')

df.loc[df['speciality'].str.contains('Commercial'), 'speciality'] = 'Sales Manager/Purchasing/Operations'
df.loc[df['speciality'].str.contains('Sales Manager'), 'speciality'] = 'Sales Manager/Purchasing/Operations'
df['speciality'] = df['speciality'].replace(['Export'], 'Sales Manager/Purchasing/Operations')
df['speciality'] = df['speciality'].replace(['Sales Manager'], 'Sales Manager/Purchasing/Operations')

df.loc[df['speciality'].str.contains('Student'), 'speciality'] = 'Student/Resident'

df.loc[df['speciality'].str.contains('Inquiries'), 'speciality'] = 'Other'
df.loc[df['speciality'].str.contains('Inquiry'), 'speciality'] = 'Other'
df.loc[df['speciality'].str.contains('Graphist'), 'speciality'] = 'Other'
df.loc[df['speciality'].str.contains('Design'), 'speciality'] = 'Other'
df.loc[df['speciality'].str.contains('Finance'), 'speciality'] = 'Other'
df['speciality'] = df['speciality'].replace(['Administration'], 'Other')
df['speciality'] = df['speciality'].replace(['Bme'], 'Other')
df['speciality'] = df['speciality'].replace(['Data Anlyse'], 'Other')
df['speciality'] = df['speciality'].replace(['Customercare'], 'Other')
df['speciality'] = df['speciality'].replace(['Connecting Celebrities To Brands'], 'Other')
df['speciality'] = df['speciality'].replace(['N/A '], 'Other')
df['speciality'] = df['speciality'].replace(['N/A'], 'Other')
df['speciality'] = df['speciality'].replace(['Data Analyse'], 'Other')
df['speciality'] = df['speciality'].replace(['Hugel'], 'Other')
df['speciality'] = df['speciality'].replace(['Go'], 'Other')
df['speciality'] = df['speciality'].replace(['Omfs'], 'Other')
df['speciality'] = df['speciality'].replace(['Non'], 'Other')
df['speciality'] = df['speciality'].replace(['No'], 'Other')
df['speciality'] = df['speciality'].replace(['-'], 'Other')
df['speciality'] = df['speciality'].replace(['All'], 'Other')
df['speciality'] = df['speciality'].replace(['Gum'], 'Other')


# DEDUPING
dfDeduped = df.drop_duplicates()
dfDeduped['JoomlaFixQuery'] = 'UPDATE joo_acymailing_subscriber SET eloqua = "Yes" WHERE email LIKE "' + dfDeduped['email'] + '" ;'

# CLEANING
ind_drop = dfDeduped[dfDeduped['email'].apply(lambda x: x.endswith('@informa.com'))].index
dfDeduped = dfDeduped.drop(ind_drop)
ind_drop = dfDeduped[dfDeduped['email'].apply(lambda x: x.endswith('@euromedicom.com'))].index
dfDeduped = dfDeduped.drop(ind_drop)
ind_drop = dfDeduped[dfDeduped['email'].apply(lambda x: x.endswith('@eurogin.com'))].index
dfDeduped = dfDeduped.drop(ind_drop)
ind_drop = dfDeduped[dfDeduped['email'].apply(lambda x: x == ('georges.hinot@gmail.com'))].index
dfDeduped = dfDeduped.drop(ind_drop)
ind_drop = dfDeduped[dfDeduped['email'].apply(lambda x: x == ('sauveneelaurent@gmail.com'))].index
dfDeduped = dfDeduped.drop(ind_drop)
ind_drop = dfDeduped[dfDeduped['email'].apply(lambda x: x == ('hakimislim@yahoo.fr'))].index
dfDeduped = dfDeduped.drop(ind_drop)

# PRINT
print(tabulate(dfDeduped, headers='keys', tablefmt='psql', showindex=False))

# COUNT
number = df.shape[0]
print('Original:', number)
numberFinal = dfDeduped.shape[0]
print('Final:', numberFinal)

# EXPORT & EXCEL
outputExcelFile = r'C:/Users/Georges/Downloads/INTEGRATION/'+myToday+'_AEM-Eloqua.xlsx'

dfDeduped.to_excel(outputExcelFile, index=False, sheet_name='AEM-Eloqua', header=['source', 'theme', 'type', 'title', 'firstname', 'lastname', 'email', 'speciality', 'country', 'phone', 'company', 'JoomlaFixQuery'])
workbook = openpyxl.load_workbook(outputExcelFile)
worksheet = workbook['AEM-Eloqua']
FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
worksheet.auto_filter.ref = FullRange
worksheet.freeze_panes = 'A2'
sheetsLits = workbook.sheetnames
workbook['AEM-Eloqua'].column_dimensions['A'].width = 30
workbook['AEM-Eloqua'].column_dimensions['B'].width = 15
workbook['AEM-Eloqua'].column_dimensions['C'].width = 15
workbook['AEM-Eloqua'].column_dimensions['D'].width = 10
workbook['AEM-Eloqua'].column_dimensions['E'].width = 15
workbook['AEM-Eloqua'].column_dimensions['F'].width = 30
workbook['AEM-Eloqua'].column_dimensions['G'].width = 30
workbook['AEM-Eloqua'].column_dimensions['H'].width = 15
workbook['AEM-Eloqua'].column_dimensions['I'].width = 15
workbook['AEM-Eloqua'].column_dimensions['J'].width = 15
workbook['AEM-Eloqua'].column_dimensions['K'].width = 15
workbook['AEM-Eloqua'].column_dimensions['L'].width = 90
for sheet in sheetsLits:
    worksheet = workbook[sheet]
    for cell in workbook[sheet][1]:
        worksheet[cell.coordinate].fill = PatternFill(fgColor = 'FFC6C1C1', fill_type = 'solid')
workbook.save(outputExcelFile)