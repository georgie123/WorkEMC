import os
from tabulate import tabulate as tab

import pandas as pd

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
from matplotlib.collections import PatchCollection
from mpl_toolkits.basemap import Basemap

import numpy as np

from PIL import Image, ImageOps

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


shp_simple_countries = r'C:/Users/Georges/PycharmProjects/data/simple_countries/simple_countries'

workDirectory = r'C:/Users/Georges/Downloads/'

segmentFileName = 'Sponsored_Teoxane'

outputExcelFile = workDirectory+segmentFileName+'_Counts.xlsx'

# Excel import
inputExcelFile = workDirectory+segmentFileName+'.xlsx'
df = pd.read_excel(inputExcelFile, sheet_name='union', engine='openpyxl',
                   usecols=['Email Address', 'Job Title', 'Company', 'City', 'State or Province', 'Country Name'])


# COUNT COUNTRY
df_Country_count = pd.DataFrame(df.groupby(['Country Name'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Country_count = df_Country_count.fillna('Unknow')

df_Country_count['Percent'] = (df_Country_count['Total'] / df_Country_count['Total'].sum()) * 100
df_Country_count['Percent'] = df_Country_count['Percent'].round(decimals=2)


# COUNT JOB TITLE
df['Job Title'] = df['Job Title'].str.replace(r'[,;.:()%]', '', regex=True)
df_Job_count = pd.DataFrame(df.groupby(['Job Title'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Job_count = df_Job_count.fillna('Unknown')

df_Job_count['Percent'] = (df_Job_count['Total'] / df_Job_count['Total'].sum()) * 100
df_Job_count['Percent'] = df_Job_count['Percent'].round(decimals=2)


# COUNT JOB TITLE PER COUNTRY
df_SpecialtiesPerCountry_count = pd.DataFrame(df.groupby(['Country Name', 'Job Title'], dropna=False)\
    .size(), columns=['Total']).sort_values(['Country Name', 'Total'], ascending=[True, False]).reset_index()
df_SpecialtiesPerCountry_count = df_SpecialtiesPerCountry_count.fillna('Unknown')

df_SpecialtiesPerCountry_count['Percent'] = (df_SpecialtiesPerCountry_count['Total'] / df_SpecialtiesPerCountry_count['Total'].sum()) * 100
df_SpecialtiesPerCountry_count['Percent'] = df_SpecialtiesPerCountry_count['Percent'].round(decimals=2)


# COUNT COMPANY
df['Company'] = df['Company'].str.upper().str.replace(r'[,;.:()%]', '', regex=True)
df['Company'] = df['Company'].str.upper().str.replace(' S R L', ' SRL').str.replace('S R L ', 'SRL ')
df_Company_count = pd.DataFrame(df.groupby(['Company'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Company_count = df_Company_count.fillna('UNKNOWN OR NONE')

df_Company_count['Percent'] = (df_Company_count['Total'] / df_Company_count['Total'].sum()) * 100
df_Company_count['Percent'] = df_Company_count['Percent'].round(decimals=2)


# COUNT EMAIL DOMAINS
df['Email Address'] = df['Email Address'].str.lower()
df['Domain'] = df['Email Address'].str.split('@').str[1]
df_Email_DNS_count = pd.DataFrame(df.groupby(['Domain'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Email_DNS_count = df_Email_DNS_count.fillna('Unknown')

df_Email_DNS_count['Percent'] = (df_Email_DNS_count['Total'] / df_Email_DNS_count['Total'].sum()) * 100
df_Email_DNS_count['Percent'] = df_Email_DNS_count['Percent'].round(decimals=2)


# COUNT STATE
df['State or Province'] = df['State or Province'].str.upper().str.replace(r'\d+', '', regex=True).str.replace(r'[,;.:()%]', '', regex=True)
df['State or Province'] = df['State or Province'].replace(r'^\s+$', np.nan, regex=True)
df_State_count = pd.DataFrame(df.groupby(['Country Name', 'State or Province'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Country Name', 'Total'], ascending=[True, False]).reset_index()
df_State_count = df_State_count.fillna('UNKNOWN')

df_State_count['Percent'] = (df_State_count['Total'] / df_State_count['Total'].sum()) * 100
df_State_count['Percent'] = df_State_count['Percent'].round(decimals=2)


# COUNT CITY
df['City'] = df['City'].str.upper().str.replace(r'\d+', '', regex=True).str.title().str.replace(r'[,;.:()%]', '', regex=True)
df['City'] = df['City'].replace(r'^\s+$', np.nan, regex=True)
df_City_count = pd.DataFrame(df.groupby(['Country Name', 'City'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Country Name', 'Total'], ascending=[True, False]).reset_index()
df_City_count = df_City_count.fillna('Unknown')

df_City_count['Percent'] = (df_City_count['Total'] / df_City_count['Total'].sum()) * 100
df_City_count['Percent'] = df_City_count['Percent'].round(decimals=2)


# EXCEL FILE
writer = pd.ExcelWriter(outputExcelFile, engine='xlsxwriter')

df_Country_count.to_excel(writer, index=False, sheet_name='Countries', header=['Country', 'Total', '%'])
df_Job_count.to_excel(writer, index=False, sheet_name='Job titles', header=['Job title', 'Total', '%'])
df_SpecialtiesPerCountry_count.to_excel(writer, index=False, sheet_name='Job titles per country', header=['Country', 'Job title', 'Total', '%'])
df_Company_count.to_excel(writer, index=False, sheet_name='Companies', header=['Company', 'Total', '%'])
df_Email_DNS_count.to_excel(writer, index=False, sheet_name='Email domains', header=['Email domain', 'Total', '%'])
df_State_count.to_excel(writer, index=False, sheet_name='States', header=['Country', 'State', 'Total', '%'])
df_City_count.to_excel(writer, index=False, sheet_name='Cities', header=['Country', 'City', 'Total', '%'])

writer.save()


# EXCEL FILTERS
workbook = openpyxl.load_workbook(outputExcelFile)
sheetsLits = workbook.sheetnames

for sheet in sheetsLits:
    worksheet = workbook[sheet]
    FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    workbook.save(outputExcelFile)

# EXCEL COLORS
for sheet in sheetsLits:
    worksheet = workbook[sheet]
    for cell in workbook[sheet][1]:
        worksheet[cell.coordinate].fill = PatternFill(fgColor = 'FFC6C1C1', fill_type = 'solid')
        workbook.save(outputExcelFile)

# EXCEL COLUMN SIZE
for sheet in sheetsLits:
    for cell in workbook[sheet][1]:
        if get_column_letter(cell.column) == 'A':
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 30
        else:
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 10
        workbook.save(outputExcelFile)


# MAP COUNTRIES
df_Country_count.set_index('Country Name', inplace=True)

my_values = df_Country_count['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_Country_count['Percent'] = np.digitize(my_values, my_range) - 1

map1 = plt.figure(figsize=(14, 8))

ax = map1.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_countries, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['CTY_ELOQUA']
    if shp_ctry not in df_Country_count.index:
        color = '#dddddd'
    else:
        color = scheme[df_Country_count.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map1.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# cb.ax.set_xticklabels([str(round(i, 1)) for i in my_range])
# cb.ax.tick_params(labelsize=7)
# cb.set_label('Percentage', rotation=0)
cb.remove()

map1.savefig(workDirectory+'mymap1.png', dpi=110, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap1.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap1.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap1.png')
img.anchor = 'E2'

workbook['Countries'].add_image(img)
workbook.save(outputExcelFile)


# REMOVE PICTURES
os.remove(workDirectory+'mymap1.png')


# TERMINAL OUTPUTS AND TESTS
print(tab(df_Country_count.head(20), headers='keys', tablefmt='psql'))
print("OK, export done!")