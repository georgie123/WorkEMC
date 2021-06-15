
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from datetime import date
import pandas as pd
from tabulate import tabulate
import numpy as np

today = date.today()

alreadyHandledTAS = [5476324,5476282,5476275,5476262,5476258,5476255,5476252,3709874,5403438,3950740,5388809,5388804,3960884,5264434,5264427,5015181,5015182,5015185,5056113,5074020,5074021,3952928,5110112,5110116,5110122,5248797,5248801,5248804,5248806,5248809,5248810,5248815,5248816,5248818,5248820,5248822,5248825]
alreadyHandledMCS = [4975694,3949499,4037270,3871032,4929940,5248828]
alreadyHandledVCS = [4943822,5476351,5476344,5476338,3834070,5411440,4017075,5403445,5403444,5403441,3952927,5388832,3979549,5264446,5113554,5006917,3880324,3846533,5095240,5095246,5095285,5095362,5095365,5095368,4994975,4967808,4974001,4967878,5056136,5056142,4994981,5056145,5056147,3995729,5074010,3880459,5074013,5074015,4057924,5110134,5110138,5110140,5110143,5110144,3887114,5110145,5110147,5110150,5110154,3834085,5110158,5110162,3979620,4074316,5248663,3710287,5248667,3834180,5248674,5248756,5248760,5248763,5248775,5248780]
alreadyHandled = alreadyHandledTAS + alreadyHandledMCS + alreadyHandledVCS

PremiumWithEmail2 = ['mcsweeney2@me.com', 'XXX']

# CAS PARTICULIER EN 2021
# 3979620 : a déjà un membership payé, j'ai ajouté "VCS 2021" et 1 an de membership, mais rien ajouté dans amount, car c'est celui en euro.
# 5000140 : a déjà 2 memberships, mais le second est déjà à 350 dollars, sans doute un ajout manuel de l'équipe américaine. J'ai ajouté "VCS 2021" et 1 an de membership (jusqu'en 2023) puisque le premier, payé, est déjà valable jusqu'en 2022.

df_FichierTas = pd.read_csv(r'C:/Users/Georges/Downloads/Georges report on on-demand_AMS purchases (TAS) .csv', sep=',', usecols=[
    'Would you like to purchase On-Demand with AMS Premium Membership?', 'First Name', 'Last Name', 'Email Address', 'CC Email Address',
    'Job Title', 'Parent Attendee: Degree', 'Parent Attendee: Medical Specialty:', 'Parent Attendee: If \'other\' please specify',
    'Parent Attendee: How many years have you been in practice?',
    'Address 1', 'Address 2', 'Zip/Postal Code', 'City', 'State/Province', 'Country',
    'Parent Attendee: Mobile Phone', 'Parent Attendee: Work Phone'])

df_FichierMcs = pd.read_csv(r'C:/Users/Georges/Downloads/Georges report on on-demand_AMS purchases (MCS) .csv', sep=',', usecols=[
    'Would you like to purchase On-Demand with AMS Premium Membership?', 'First Name', 'Last Name', 'Email Address', 'CC Email Address',
    'Practice Type', 'Degree', 'Please specify:',
    'Parent Attendee: How many years have you been in practice?',
    'Address 1', 'Address 2', 'Zip/Postal Code', 'City', 'State/Province', 'Country'])

df_FichierVcs = pd.read_excel(r'C:/Users/Georges/Downloads/Georges report on on-demand_AMS purchases (VCS).xlsx',
                   sheet_name='Georges report on on-demandAMS', engine='openpyxl', usecols=[
        'Wednesday, June 23, 2021: Would you like to purchase On-Demand with...', 'First Name', 'Last Name', 'Email Address',
        'Job Title',
        'Address 1', 'Address 2', 'Zip/Postal Code', 'City', 'State/Province', 'Country',
        'Work Phone'])

df_AMS = pd.read_excel(r'C:/Users/Georges/Downloads/User_export_'+str(today)+'.xlsx', sheet_name='Export', engine='openpyxl',
                   usecols=['ID', 'Email', 'Last Membership:Type name', 'Last Membership:Expires at'
                            ])


# ADD MISSING FIELD TAS
df_FichierTas['source'] = 'TAS 2021'

# ADD MISSING FIELD MCS
df_FichierMcs['source'] = 'MCS 2021'
df_FichierMcs['Parent Attendee: Mobile Phone'] = np.NaN
df_FichierMcs['Parent Attendee: Work Phone'] = np.NaN
df_FichierMcs['Parent Attendee: If \'other\' please specify'] = np.NaN

# ADD MISSING FIELD VCS
df_FichierVcs['CC Email Address'] = np.NaN
df_FichierVcs['source'] = 'VCS 2021'
df_FichierVcs['Parent Attendee: Mobile Phone'] = np.NaN
df_FichierVcs['Parent Attendee: Degree'] = np.NaN
df_FichierVcs['Parent Attendee: Medical Specialty:'] = np.NaN
df_FichierVcs['Parent Attendee: If \'other\' please specify'] = np.NaN
df_FichierVcs['Parent Attendee: How many years have you been in practice?'] = np.NaN


# Rename fields from TAS
df_FichierTas.rename(columns={'Would you like to purchase On-Demand with AMS Premium Membership?': 'AMS Premium',
                              'Parent Attendee: Mobile Phone': 'Mobile', 'Parent Attendee: Work Phone': 'Work Phone',
                              'Job Title': 'Spe1', 'Parent Attendee: Degree': 'Spe2', 'Parent Attendee: Medical Specialty:': 'Spe3', 'Parent Attendee: If \'other\' please specify': 'Spe4',
                              'Parent Attendee: How many years have you been in practice?': 'Experience'}, inplace=True)
# Rename fields from MCS
df_FichierMcs.rename(columns={'Would you like to purchase On-Demand with AMS Premium Membership?': 'AMS Premium',
                              'Parent Attendee: Mobile Phone': 'Mobile', 'Parent Attendee: Work Phone': 'Work Phone',
                              'Practice Type': 'Spe1', 'Degree': 'Spe2', 'Please specify:': 'Spe3', 'Parent Attendee: If \'other\' please specify': 'Spe4',
                              'Parent Attendee: How many years have you been in practice?': 'Experience'}, inplace=True)
# Rename fields from VCS
df_FichierVcs.rename(columns={'Wednesday, June 23, 2021: Would you like to purchase On-Demand with...': 'AMS Premium',
                              'Parent Attendee: Mobile Phone': 'Mobile', 'Parent Attendee: Work Phone': 'Work Phone',
                              'Job Title': 'Spe1', 'Parent Attendee: Degree': 'Spe2', 'Parent Attendee: Medical Specialty:': 'Spe3', 'Parent Attendee: If \'other\' please specify': 'Spe4',
                              'Parent Attendee: How many years have you been in practice?': 'Experience'}, inplace=True)


TAS = df_FichierTas.loc[df_FichierTas['AMS Premium'] == 'Yes, I would like to purchase On-Demand with AMS Premium Membership']
MCS = df_FichierMcs.loc[df_FichierMcs['AMS Premium'] == 'Yes, I would like to purchase On-Demand with AMS Premium Membership']
VCS = df_FichierVcs.loc[df_FichierVcs['AMS Premium'] == 'Yes,  I would like to purchase On-Demand with AMS Premium Membership']


# Merge the selections to create the Premiums from Aventri
df_Merged = pd.concat([TAS, MCS, VCS], axis=0)


# Fix fields in Premium from Aventri
df_Merged['First Name'] = df_Merged['First Name'].str.title()
df_Merged['Last Name'] = df_Merged['Last Name'].str.title()
df_Merged['Email Address'] = df_Merged['Email Address'].str.lower()
df_Merged['CC Email Address'] = df_Merged['CC Email Address'].str.lower()
df_Merged.loc[df_Merged['CC Email Address'] == df_Merged['Email Address'], 'CC Email Address'] = np.NaN

df_Merged['Live location'] = df_Merged['Address 1'].fillna('') + ' ' + df_Merged['Address 2'].fillna('') \
                             + ' ' + df_Merged['Zip/Postal Code'].fillna('') + ' ' + df_Merged['City'].fillna('') + ' ' + df_Merged['State/Province'].fillna('') \
                             + ' ' + df_Merged['Country'].fillna('')

df_Merged['Specialty'] = df_Merged['Spe1'].fillna('') + ' ' + df_Merged['Spe2'].fillna('') \
                             + ' ' + df_Merged['Spe3'].fillna('') + ' ' + df_Merged['Spe4'].fillna('')

# Re-Organize Premium from Aventri
df_Merged = df_Merged[['source', 'First Name', 'Last Name', 'Email Address', 'CC Email Address', 'Specialty', 'Experience',
                       'Live location',
                       'Mobile', 'Work Phone']]


# Find duplicates in Premium from Aventri
duplicatePremiumAventri = df_Merged[df_Merged.duplicated(['Email Address'])]
duplicatePremiumAventri = duplicatePremiumAventri[['Email Address', 'source']]


# Fix fields from AMS
df_AMS['Email'] = df_AMS['Email'].str.lower()


# ARE IN AVENTRI FOR PREMIUM BUT ARE NOT IN AMS
df_difference = pd.DataFrame(df_Merged[~df_Merged['Email Address'].isin(df_AMS['Email'])])
df_difference = df_difference.loc[df_difference['Email Address'].isin(PremiumWithEmail2) == False]


# ARE IN AMS AND IN AVENTRI FOR PREMIUM BUT NEVER HANDLED
df_BothButNoHandled = pd.merge(df_AMS, df_Merged, left_on='Email', right_on='Email Address')
df_BothButNoHandled = df_BothButNoHandled.loc[df_BothButNoHandled['ID'].isin(alreadyHandled) == False]


print('\nAll Premiums from Aventri:\n', tabulate(df_Merged.head(10), headers='keys', tablefmt='psql', showindex=False))
number = df_Merged.shape[0]
print('Total Premiums from Aventri:', number)
CountTAS = df_Merged[df_Merged['source'] == 'TAS 2021'].count()['source']
print('From TAS:', CountTAS)
CountMCS = df_Merged[df_Merged['source'] == 'MCS 2021'].count()['source']
print('From MCS:', CountMCS)
CountVCS = df_Merged[df_Merged['source'] == 'VCS 2021'].count()['source']
print('From VCS:', CountVCS)
numberDuplicates = duplicatePremiumAventri.shape[0]
print('Duplicates ('+str(numberDuplicates)+'):\n', tabulate(duplicatePremiumAventri, headers=(), tablefmt='plain', showindex=False))


print('\nAll AMS:\n', tabulate(df_AMS.head(10), headers='keys', tablefmt='psql', showindex=False))
numberAMS = df_AMS.shape[0]
print('Total from AMS:', numberAMS)

print('\nPremiums from Aventri to add:\n', tabulate(df_difference.head(10).fillna(''), headers='keys', tablefmt='psql', showindex=False))
numberDiff = df_difference.shape[0]
print('Total to add in AMS:', numberDiff)

print('\nPremiums from Aventri already in AMS but never handled:\n', tabulate(df_BothButNoHandled.head(10).fillna(''), headers='keys', tablefmt='psql', showindex=False))
numberBothButNoHandled = df_BothButNoHandled.shape[0]
print('Total to handled in AMS:', numberBothButNoHandled)


# Export Premium from Aventri
FieldsPremiumFromAventri = ['source', 'First Name', 'Last Name', 'Email Address', 'CC Email Address', 'Specialty', 'Experience',
                       'Live location',
                       'Mobile', 'Work Phone']

PremiumFromAventri = r'C:/Users/Georges/Downloads/'+str(today)+' Premium from Aventri.xlsx'
df_Merged.to_excel(PremiumFromAventri, index=False, sheet_name='Premium From Aventri', header=FieldsPremiumFromAventri)

workbook = openpyxl.load_workbook(PremiumFromAventri)

worksheet = workbook['Premium From Aventri']
FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
worksheet.auto_filter.ref = FullRange
sheetsLits = workbook.sheetnames
workbook['Premium From Aventri'].column_dimensions['A'].width = 15
workbook['Premium From Aventri'].column_dimensions['B'].width = 15
workbook['Premium From Aventri'].column_dimensions['C'].width = 15
workbook['Premium From Aventri'].column_dimensions['D'].width = 30
workbook['Premium From Aventri'].column_dimensions['E'].width = 30
workbook['Premium From Aventri'].column_dimensions['F'].width = 20
workbook['Premium From Aventri'].column_dimensions['G'].width = 20
workbook['Premium From Aventri'].column_dimensions['H'].width = 40
workbook['Premium From Aventri'].column_dimensions['I'].width = 20
workbook['Premium From Aventri'].column_dimensions['J'].width = 20
worksheet.freeze_panes = 'A2'
for sheet in sheetsLits:
    worksheet = workbook[sheet]
    for cell in workbook[sheet][1]:
        worksheet[cell.coordinate].fill = PatternFill(fgColor='FFC6C1C1', fill_type='solid')
workbook.save(PremiumFromAventri)